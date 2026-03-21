import pandas as pd
import openpyxl
import os

# Archivos solicitados
f1 = r"c:\Users\Facun\OneDrive\Escritorio\PROYECTOS PERSONALES\PRECIOS\Copia de VITAL2-nini-products-20250921.xlsx"
f2 = r"c:\Users\Facun\OneDrive\Escritorio\PROYECTOS PERSONALES\PRECIOS\VITAL-all-products-20250921.xlsx"

def analyze_file(path, label, memory_safe=False):
    print(f"\n--- ANALIZANDO: {label} ---")
    if not os.path.exists(path):
        print(f"⚠️ Archivo no encontrado: {path}")
        return

    try:
        if memory_safe:
            df_sample = pd.read_excel(path, nrows=10)
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            rows, cols = ws.max_row, ws.max_column
            wb.close()
        else:
            df_all = pd.read_excel(path)
            rows, cols = df_all.shape
            df_sample = df_all.head(5)

        print(f"1. Dimensiones: {rows} filas x {cols} columnas")
        cols_list = df_sample.columns.tolist()
        print(f"2. Nombres de columnas: {cols_list}")
        
        ean_cols = [c for c in cols_list if any(kw in c.lower() for kw in ["ean", "bar", "cod", "sku"])]
        print(f"3. Referencia Código (EAN/SKU): {'SÍ' if ean_cols else 'NO'} ({ean_cols if ean_cols else '-'})")
        
        comp_cols = [c for c in cols_list if any(kw in c.lower() for kw in ["site", "compet", "shop", "fuent", "orig"])]
        print(f"4. Información de competidores: {'SÍ' if comp_cols else 'NO'} ({comp_cols if comp_cols else '-'})")
        
        print("5. Ejemplo de filas:")
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', 1000)
        print(df_sample.head(3))
        
    except Exception as e:
        print(f"❌ Error analizando {label}: {e}")

if __name__ == "__main__":
    analyze_file(f1, "VITAL2-nini (987 KB)")
    analyze_file(f2, "VITAL-all (91 MB)", memory_safe=True)
