import os
import json
import openpyxl
import pandas as pd

json_path = r"c:\Users\Facun\OneDrive\Escritorio\PROYECTOS PERSONALES\PRECIOS\BRUJULA-DE-PRECIOS\data\output_maxiconsumo.json"
excel_path = r"c:\Users\Facun\OneDrive\Escritorio\PROYECTOS PERSONALES\PRECIOS\VITAL-all-products-20250921.xlsx"

def clean_ean(ean):
    if ean is None: return None
    s = str(ean).split('.')[0].strip()
    return s if s and s != 'nan' and s != 'None' else None

def do_match():
    # 1. Cargar JSON
    if not os.path.exists(json_path):
        print("❌ Error: No existe output_maxiconsumo.json")
        return
        
    with open(json_path, "r", encoding="utf-8") as f:
        catalog = json.load(f)
        
    skus = {str(item.get("sku", "")).strip() for item in catalog if item.get("sku")}
    print(f"📊 Nuestro catálogo actual tiene {len(skus):,} productos únicos.")

    # 2. Cargar Excel usando openpyxl en modo read_only (mucho más rápido para 91MB)
    print(f"⌛ Extrayendo EANs del Excel de 91MB (Modo Read-Only)...")
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Identificar columna EAN
        ean_idx = -1
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for i, col_name in enumerate(header):
            if col_name and str(col_name).lower() == 'ean':
                ean_idx = i
                break
        
        if ean_idx == -1:
            print("❌ Error: No se encontró la columna 'ean' en el Excel.")
            return

        excel_eans = set()
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            val = row[ean_idx]
            clean = clean_ean(val)
            if clean:
                excel_eans.add(clean)
            if i % 100000 == 0:
                print(f"   Processed {i:,} rows...")
        
        wb.close()
        
        # 3. Cruzar
        coincidencias = skus.intersection(excel_eans)
        
        print("\n--- RESULTADO DEL CRUCE ---")
        print(f"✅ Coincidencias encontradas: {len(coincidencias):,}")
        print(f"📈 Porcentaje de integración: {(len(coincidencias)/len(skus)*100):.1f}%")
        
    except Exception as e:
        print(f"❌ Error al procesar: {e}")

if __name__ == "__main__":
    do_match()
