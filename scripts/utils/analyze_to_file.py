import pandas as pd
import os

f1 = r"c:\Users\Facun\OneDrive\Escritorio\PROYECTOS PERSONALES\PRECIOS\Copia de VITAL2-nini-products-20250921.xlsx"
f2 = r"c:\Users\Facun\OneDrive\Escritorio\PROYECTOS PERSONALES\PRECIOS\VITAL-all-products-20250921.xlsx"

def analyze_and_file(path, label, prefix):
    print(f"\n--- {label} ---")
    try:
        df = pd.read_excel(path, nrows=10)
        cols = df.columns.tolist()
        with open(f'{prefix}_cols.txt', 'w') as f:
            f.write("\n".join(cols))
        
        # Guardamos el head en un CSV para leerlo mas facil
        df.head(3).to_csv(f'{prefix}_sample.csv', index=False)
        
        # Dimensiones
        if label == "VITAL-all (91 MB)":
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            rows, cols_cnt = ws.max_row, ws.max_column
            wb.close()
        else:
            rows, cols_cnt = pd.read_excel(path).shape
        
        print(f"DONE: {rows} x {cols_cnt}")
    except Exception as e:
        print("ERROR:", e)

analyze_and_file(f1, "VITAL2-nini (987 KB)", "v_nini")
analyze_and_file(f2, "VITAL-all (91 MB)", "v_all")
