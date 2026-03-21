import os
import json
import openpyxl

json_path = r"c:\Users\Facun\OneDrive\Escritorio\PROYECTOS PERSONALES\PRECIOS\BRUJULA-DE-PRECIOS\data\output_maxiconsumo.json"
excel_path = r"c:\Users\Facun\OneDrive\Escritorio\PROYECTOS PERSONALES\PRECIOS\VITAL-all-products-20250921.xlsx"

def clean_ean(ean):
    if ean is None: return None
    s = str(ean).split('.')[0].strip()
    return s if s and s != 'nan' and s != 'None' else None

def do_match():
    with open(json_path, "r", encoding="utf-8") as f:
        catalog = json.load(f)
    skus = {str(item.get("sku", "")).strip() for item in catalog if item.get("sku")}
    
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    ean_idx = -1
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for i, col_name in enumerate(header):
        if col_name and str(col_name).lower() == 'ean':
            ean_idx = i
            break
            
    excel_eans = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[ean_idx]
        c = clean_ean(val)
        if c: excel_eans.add(c)
    
    matches = skus.intersection(excel_eans)
    print(f"COINCIDENCIAS ENCONTRADAS ({len(matches)}): {matches}")
    
    # Mostrar un ejemplo de por qué falló la mayoría
    print("\nEJEMPLO DE NUESTROS SKUS (Primeros 5):", list(skus)[:5])
    print("EJEMPLO DE EANs EN EXCEL (Primeros 5 de la columna 'ean'):", list(excel_eans)[:5])

if __name__ == "__main__":
    do_match()
