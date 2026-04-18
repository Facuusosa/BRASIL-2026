#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UNIFICADOR V2 — Brujula de Precios
Misma logica que actualizar_catalogo.py con tres mejoras:
  1. confianza_match por producto: exacto_ean | fuzzy_alto | fuzzy
  2. Threshold fuzzy subido: 0.60 -> 0.75 (menos falsos positivos)
  3. Reporte detallado al final
"""

import os, json, glob, re, unicodedata
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    print("[WARN] pip install openpyxl")

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
YAGUAR_DIR      = os.path.join(BASE_DIR, "targets", "yaguar")
MAXICARRE_DIR   = os.path.join(BASE_DIR, "targets", "maxicarrefour")
MAXICONSUMO_DIR = os.path.join(BASE_DIR, "targets", "maxiconsumo")
CODIGOS_FILE    = os.path.join(BASE_DIR, "data", "raw", "CODIGOS.xlsx")
MAESTRO_FILE    = os.path.join(BASE_DIR, "data", "raw", "Listado Maestro 09-03.xlsx")
OUTPUT_FILE     = os.path.join(BASE_DIR, "BRUJULA-DE-PRECIOS", "data", "processed", "catalogo_unificado.json")

_FUZZ_TH = 0.75  # subido desde 0.60

SECTOR_MAP = {
    "almacen": "Almacén", "almacén": "Almacén",
    "bebidas": "Bebidas", "bodega": "Bebidas",
    "frescos": "Frescos", "lacteos": "Frescos", "lácteos": "Frescos",
    "lácteos y productos frescos": "Frescos", "lacteos y productos frescos": "Frescos",
    "fiambreria": "Frescos", "fiambrería": "Frescos",
    "carniceria": "Frescos", "carnicería": "Frescos", "quesos": "Frescos",
    "limpieza": "Limpieza", "papeles e higiene": "Limpieza", "papeles": "Limpieza",
    "perfumeria": "Cuidado Personal", "perfumería": "Cuidado Personal",
    "cuidado personal": "Cuidado Personal",
    "bazar": "Bazar", "bazar y textil": "Bazar", "hogar y bazar": "Bazar",
    "mascotas": "Mascotas", "kiosco": "Kiosco",
    "mundo bebe": "Bebés", "mundo bebé": "Bebés", "bebes": "Bebés", "bebés": "Bebés",
    "congelados": "Congelados",
    "desayuno y merienda": "Desayuno y Merienda", "desayuno": "Desayuno y Merienda",
    "panaderia": "Almacén", "panadería": "Almacén",
}

def norm_sector(raw):
    return SECTOR_MAP.get((raw or "").lower().strip(), (raw or "Almacén").strip().title())

def norm_nombre(nombre):
    n = (nombre or "").lower().strip()
    n = unicodedata.normalize("NFD", n)
    n = "".join(c for c in n if unicodedata.category(c) != "Mn")
    n = re.sub(r"(\d),(\d)", r"\1.\2", n)
    n = re.sub(r"[^a-z0-9. ]", " ", n)
    n = re.sub(r"\bx\s*(\d)", r"\1", n)
    n = re.sub(r"(\d+\.?\d*)\s*lts?\b", lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    n = re.sub(r"(\d+\.?\d*)\s*lt\b",   lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    n = re.sub(r"(\d+\.?\d*)\s*l\b",    lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    n = re.sub(r"(\d+)\s*grs\b", lambda m: m.group(1)+"gr", n)
    n = re.sub(r"(\d+\.?\d*)\s*kgs?\b", lambda m: str(int(float(m.group(1))*1000))+"gr", n)
    n = re.sub(r"(\d+)\s*uni\b", lambda m: m.group(1)+"un", n)
    n = re.sub(r"(\d+)\s*(cc|ml|gr|kg|un)\b", r"\1\2", n)
    n = re.sub(r"\.", " ", n)
    return re.sub(r"\s+", " ", n).strip()

def display_nombre(nombre):
    n = (nombre or "").strip()
    n = re.sub(r"\bX\s*(?=\d)", "", n)
    minusc = {"de","del","la","las","los","el","y","e","o","a","con","sin","en","al","por"}
    palabras = n.split()
    tc = []
    for i, p in enumerate(palabras):
        if not p: continue
        if i > 0 and p.lower() in minusc:
            tc.append(p.lower())
        else:
            tc.append(p[0].upper() + p[1:] if len(p) > 1 else p.upper())
    n = " ".join(tc)
    n = re.sub(r"(\d)\s*[Cc][Cc]\b", r"\1 ml", n)
    n = re.sub(r"(\d)\s*[Ll][Tt][Ss]?\b", r"\1 L", n)
    n = re.sub(r"(\d)\s*[Mm][Ll][Ss]?\b", r"\1 ml", n)
    n = re.sub(r"(\d)\s*[Gg][Rr][Ss]?\b", r"\1 g", n)
    n = re.sub(r"(\d)\s*[Kk][Gg][Ss]?\b", r"\1 kg", n)
    return re.sub(r"\s+", " ", n).strip().strip(".")

def mejor_imagen(imgs):
    for i in imgs:
        if i and "/0000-" not in i: return i
    return next((i for i in imgs if i), "")

def precio_valido(p): return p > 200

# ---------------------------------------------------------------------------
def cargar_excel():
    yag_sku_ean, mco_sku_ean = {}, {}
    ean_yag_sku, ean_mco_sku = {}, {}
    ean_master, nombre_ean   = {}, {}

    if not EXCEL_OK:
        return yag_sku_ean, mco_sku_ean, ean_yag_sku, ean_mco_sku, ean_master, nombre_ean

    if os.path.isfile(CODIGOS_FILE):
        wb = openpyxl.load_workbook(CODIGOS_FILE, read_only=True, data_only=True)
        for row in wb["YAGUAR"].iter_rows(min_row=2, values_only=True):
            try:
                sku, ean = str(int(row[1])), str(int(row[2]))
                if len(ean) >= 8:
                    yag_sku_ean[sku] = ean
                    ean_yag_sku[ean] = sku
            except: pass
        for row in wb["MAXICONSUMO"].iter_rows(min_row=2, values_only=True):
            try:
                sku, ean = str(int(row[1])), str(int(row[3]))
                if len(ean) >= 8:
                    mco_sku_ean[sku] = ean
                    ean_mco_sku[ean] = sku
            except: pass
        wb.close()
        print(f"  CODIGOS: Yaguar={len(yag_sku_ean)} SKUs, Maxiconsumo={len(mco_sku_ean)} SKUs")

    if os.path.isfile(MAESTRO_FILE):
        wb = openpyxl.load_workbook(MAESTRO_FILE, read_only=True, data_only=True)
        for row in wb["Sheet1"].iter_rows(min_row=2, values_only=True):
            nombre, abc, sector, ean_col, barcode = row[1], row[2], row[3], row[6], row[8]
            ean_val = None
            for v in (ean_col, barcode):
                if v and str(v).strip() not in ("-", "", "None"):
                    try: ean_val = str(int(float(str(v)))); break
                    except: pass
            if not ean_val or not nombre: continue
            nombre_str = str(nombre).strip()
            ean_master[ean_val] = {
                "nombre": display_nombre(nombre_str),
                "sector": norm_sector(str(sector or "")),
                "abc":    str(abc or "").strip().upper(),
            }
            clave = norm_nombre(nombre_str)
            if clave and len(clave) > 5:
                nombre_ean[clave] = ean_val
        wb.close()
        print(f"  Maestro: {len(ean_master)} EANs, {len(nombre_ean)} nombres indexados")

    return yag_sku_ean, mco_sku_ean, ean_yag_sku, ean_mco_sku, ean_master, nombre_ean

# ---------------------------------------------------------------------------
def cargar_yaguar():
    archivos = sorted(glob.glob(os.path.join(YAGUAR_DIR, "output_yaguar_*.json")),
                      key=os.path.getmtime, reverse=True)[:8]
    sku_mejor = {}
    for f in archivos:
        try:
            data = json.load(open(f, encoding="utf-8"))
            if not data: continue
            prom = sum(p.get("precio",0) for p in data if p.get("precio",0) > 0) / max(sum(1 for p in data if p.get("precio",0) > 0), 1)
            for p in data:
                if 0 < p.get("precio",0) < 200 and prom < 200:
                    p["precio"] = round(p["precio"] * 1000, 2)
                sku = str(p.get("sku","")).strip()
                if sku and p.get("precio",0) > 0 and sku not in sku_mejor:
                    sku_mejor[sku] = p
        except: pass
    result = list(sku_mejor.values())
    print(f"  Yaguar: {len(result)} prods únicos")
    return result

def cargar_maxicarrefour():
    archivos = sorted(glob.glob(os.path.join(MAXICARRE_DIR, "output_maxicarrefour_*.json")),
                      key=os.path.getmtime, reverse=True)
    for f in archivos:
        try:
            data = json.load(open(f, encoding="utf-8"))
            con_precio = sum(1 for p in data if p.get("precio",0) > 0)
            if con_precio > 100:
                print(f"  MaxiCarrefour: {len(data)} productos ({con_precio} con precio)")
                if con_precio == 0:
                    print("  [WARN] Cookies probablemente vencidas")
                return data
        except: pass
    print("  [SKIP] No se encontró output de MaxiCarrefour")
    return []

def cargar_maxiconsumo():
    archivos = sorted(glob.glob(os.path.join(MAXICONSUMO_DIR, "output_maxiconsumo_*.json")),
                      key=os.path.getmtime, reverse=True)[:8]
    sku_mejor = {}
    for f in archivos:
        try:
            data = json.load(open(f, encoding="utf-8"))
            if not data: continue
            prom = sum(p.get("precio",0) for p in data if p.get("precio",0) > 0) / max(sum(1 for p in data if p.get("precio",0) > 0), 1)
            for p in data:
                if 0 < p.get("precio",0) < 200 and prom < 200:
                    p["precio"] = round(p["precio"] * 1000, 2)
                sku = str(p.get("sku","")).strip()
                if sku and p.get("precio",0) > 0 and sku not in sku_mejor:
                    sku_mejor[sku] = p
        except: pass
    result = list(sku_mejor.values())
    print(f"  Maxiconsumo: {len(result)} prods únicos")
    return result

# ---------------------------------------------------------------------------
def construir_catalogo(yag_data, cf_data, mco_data,
                       yag_sku_ean, mco_sku_ean,
                       ean_yag_sku, ean_mco_sku,
                       ean_master, nombre_ean):

    # Índice fuzzy por nombre sobre el Maestro
    _STOP = {"de","la","el","en","y","x","con","por","para","un","una","del","los","las","al","ml","gr","cc","kg"}
    fuzz_entries = []
    fuzz_idx = defaultdict(list)
    for clave, ean in nombre_ean.items():
        ws = {w for w in clave.split() if len(w) > 1 and w not in _STOP}
        if not ws: continue
        i = len(fuzz_entries)
        fuzz_entries.append((ws, ean))
        for w in ws: fuzz_idx[w].append(i)

    def fuzzy_ean(nombre_prod):
        cl = norm_nombre(nombre_prod)
        ws_p = {w for w in cl.split() if len(w) > 1 and w not in _STOP}
        if not ws_p: return "", 0.0
        cands = set()
        for w in ws_p:
            for i in fuzz_idx.get(w, []): cands.add(i)
        best_sim, best_ean = 0.0, ""
        for i in cands:
            ws_m, ean = fuzz_entries[i]
            inter = len(ws_p & ws_m)
            union = len(ws_p | ws_m)
            sim = inter / union if union else 0.0
            if sim > best_sim:
                best_sim, best_ean = sim, ean
        return (best_ean, best_sim) if best_sim >= _FUZZ_TH else ("", best_sim)

    def resolver_ean(sku, sku_map, nombre):
        ean = sku_map.get(str(sku).strip(), "")
        if not ean:
            ean = nombre_ean.get(norm_nombre(nombre), "")
        return ean

    def master_info(ean, fb_nombre, fb_sector):
        if ean and ean in ean_master:
            m = ean_master[ean]
            return m["nombre"], m["sector"], m.get("abc", "")
        return display_nombre(fb_nombre), norm_sector(fb_sector), ""

    def nuevo(pid, ean, nombre, imagen, sector, subcategoria, abc, confianza):
        return {
            "id_unificado": pid, "ean": ean,
            "nombre_display": nombre, "imagen": imagen,
            "sector": sector, "subcategoria": subcategoria,
            "abc": abc, "confianza_match": confianza,
            "precios": {"yaguar": 0, "maxicarrefour": 0, "maxiconsumo": 0},
            "fuentes": {},
        }

    cat = {}
    yag_by_sku = {str(p.get("sku","")).strip(): p for p in yag_data if p.get("sku")}
    mco_by_sku = {str(p.get("sku","")).strip(): p for p in mco_data if p.get("sku") and p.get("precio",0) > 0}
    yag_merged = set()
    mco_merged = set()

    # Paso 1b: enriquecer mapas inversos con Maestro fuzzy
    _FUZZ1B_TH = 0.65
    ean_yag_nuevos = ean_mco_nuevos = 0
    yag_sku_set = set(ean_yag_sku.values())
    mco_sku_set = set(ean_mco_sku.values())

    for p in yag_data:
        sku = str(p.get("sku","")).strip()
        if not sku or sku in yag_sku_set: continue
        ean_r = nombre_ean.get(norm_nombre(p.get("nombre","")), "")
        if not ean_r:
            cl = norm_nombre(p.get("nombre",""))
            ws_p = {w for w in cl.split() if len(w) > 1 and w not in _STOP}
            cands = set()
            for w in ws_p:
                for i in fuzz_idx.get(w, []): cands.add(i)
            best_sim, best_ean = 0.0, ""
            for i in cands:
                ws_m, ean = fuzz_entries[i]
                inter = len(ws_p & ws_m); union = len(ws_p | ws_m)
                sim = inter / union if union else 0.0
                if sim > best_sim: best_sim, best_ean = sim, ean
            if best_sim >= _FUZZ1B_TH: ean_r = best_ean
        if ean_r and ean_r not in ean_yag_sku:
            ean_yag_sku[ean_r] = sku; yag_sku_set.add(sku); ean_yag_nuevos += 1

    for p in mco_data:
        sku = str(p.get("sku","")).strip()
        if not sku or sku in mco_sku_set: continue
        ean_r = nombre_ean.get(norm_nombre(p.get("nombre","")), "")
        if not ean_r:
            cl = norm_nombre(p.get("nombre",""))
            ws_p = {w for w in cl.split() if len(w) > 1 and w not in _STOP}
            cands = set()
            for w in ws_p:
                for i in fuzz_idx.get(w, []): cands.add(i)
            best_sim, best_ean = 0.0, ""
            for i in cands:
                ws_m, ean = fuzz_entries[i]
                inter = len(ws_p & ws_m); union = len(ws_p | ws_m)
                sim = inter / union if union else 0.0
                if sim > best_sim: best_sim, best_ean = sim, ean
            if best_sim >= _FUZZ1B_TH: ean_r = best_ean
        if ean_r and ean_r not in ean_mco_sku:
            ean_mco_sku[ean_r] = sku; mco_sku_set.add(sku); ean_mco_nuevos += 1

    print(f"  Paso 1b: +{ean_yag_nuevos} EANs Yaguar, +{ean_mco_nuevos} EANs Maxiconsumo via Maestro")

    # Paso 2: MaxiCarrefour como hub
    stats = {"cf_yag": 0, "cf_mco": 0, "cf_ambos": 0}
    for p in cf_data:
        ean = str(p.get("ean","")).strip()
        if not ean: continue
        nombre, sector, abc = master_info(ean, p.get("nombre",""), p.get("sector",""))
        entry = nuevo(ean, ean, nombre, p.get("imagen",""), sector, p.get("subcategoria",""), abc, "exacto_ean")
        entry["precios"]["maxicarrefour"] = p.get("precio",0)
        entry["fuentes"]["maxicarrefour"] = {"nombre": p.get("nombre",""), "imagen": p.get("imagen","")}

        yag_sku = ean_yag_sku.get(ean)
        if yag_sku and yag_sku in yag_by_sku:
            yp = yag_by_sku[yag_sku]
            entry["precios"]["yaguar"] = yp.get("precio",0)
            entry["fuentes"]["yaguar"] = {"nombre": yp.get("nombre",""), "imagen": yp.get("imagen",""), "sku": yag_sku}
            yag_merged.add(yag_sku)
            stats["cf_yag"] += 1

        mco_sku = ean_mco_sku.get(ean)
        if mco_sku and mco_sku in mco_by_sku:
            mp = mco_by_sku[mco_sku]
            entry["precios"]["maxiconsumo"] = mp.get("precio",0)
            entry["fuentes"]["maxiconsumo"] = {"nombre": mp.get("nombre",""), "imagen": mp.get("imagen",""), "sku": mco_sku}
            mco_merged.add(mco_sku)
            stats["cf_mco"] += 1

        if yag_sku in yag_merged and mco_sku in mco_merged: stats["cf_ambos"] += 1

        imgs = [p.get("imagen","")]
        if yag_sku and yag_sku in yag_by_sku: imgs.append(yag_by_sku[yag_sku].get("imagen",""))
        if mco_sku and mco_sku in mco_by_sku: imgs.append(mco_by_sku[mco_sku].get("imagen",""))
        entry["imagen"] = mejor_imagen(imgs) or f"https://tupedido.carrefour.com.ar/imagenesPDA/{ean}.jpg"

        cat[ean] = entry

    print(f"  MaxiCarrefour: {len(cat)} procesados → Yaguar={stats['cf_yag']} Maxiconsumo={stats['cf_mco']} Ambos={stats['cf_ambos']}")

    # Paso 3: Yaguar no mergeado
    yag_nuevos = yag_fuzzy = 0
    for p in yag_data:
        sku = str(p.get("sku","")).strip()
        if not sku or sku in yag_merged: continue
        nombre = p.get("nombre",""); precio = p.get("precio",0)
        if not nombre: continue

        ean = resolver_ean(sku, yag_sku_ean, nombre)
        confianza = "exacto_ean" if ean else ""

        if not ean:
            ean_f, sim = fuzzy_ean(nombre)
            if ean_f:
                ean = ean_f
                confianza = "fuzzy_alto" if sim >= 0.85 else "fuzzy"
                yag_fuzzy += 1

        nombre_d, sector, abc = master_info(ean, nombre, p.get("categoria",""))
        pid = ean if ean else f"yaguar_{sku}"

        if pid not in cat:
            entry = nuevo(pid, ean, nombre_d, p.get("imagen",""), sector, "", abc, confianza or "unico")
            cat[pid] = entry
            yag_nuevos += 1
        else:
            if not cat[pid].get("confianza_match"):
                cat[pid]["confianza_match"] = confianza or "unico"

        cat[pid]["precios"]["yaguar"] = precio
        cat[pid]["fuentes"]["yaguar"] = {"nombre": nombre, "imagen": p.get("imagen",""), "sku": sku}

    print(f"  Yaguar nuevos: {yag_nuevos} ({yag_fuzzy} via fuzzy)")

    # Paso 4: Maxiconsumo no mergeado
    mco_nuevos = mco_fuzzy = 0
    for p in mco_data:
        sku = str(p.get("sku","")).strip()
        if not sku or sku in mco_merged: continue
        nombre = p.get("nombre",""); precio = p.get("precio",0)
        if not nombre: continue

        ean = resolver_ean(sku, mco_sku_ean, nombre)
        confianza = "exacto_ean" if ean else ""

        if not ean:
            ean_f, sim = fuzzy_ean(nombre)
            if ean_f:
                ean = ean_f
                confianza = "fuzzy_alto" if sim >= 0.85 else "fuzzy"
                mco_fuzzy += 1

        nombre_d, sector, abc = master_info(ean, nombre, p.get("sector",""))
        pid = ean if ean else f"maxiconsumo_{sku}"

        if pid not in cat:
            entry = nuevo(pid, ean, nombre_d, p.get("imagen",""), sector, p.get("subcategoria",""), abc, confianza or "unico")
            cat[pid] = entry
            mco_nuevos += 1
        elif cat[pid]["precios"]["maxiconsumo"] == 0:
            if confianza: cat[pid]["confianza_match"] = confianza

        cat[pid]["precios"]["maxiconsumo"] = precio
        cat[pid]["fuentes"]["maxiconsumo"] = {"nombre": nombre, "imagen": p.get("imagen",""), "sku": sku}

    print(f"  Maxiconsumo nuevos: {mco_nuevos} ({mco_fuzzy} via fuzzy)")

    # Paso 5: Validación de precios — aplica a TODOS los matches (EAN y fuzzy)
    # Si 2 precios difieren >3x, uno es un precio roto del scraper
    _MAX_RATIO = 3.0
    auditoria = []
    matches_invalidos = 0

    for entry in cat.values():
        precios_ok = {k: v for k, v in entry["precios"].items() if v > 0}
        if len(precios_ok) < 2: continue

        vals = list(precios_ok.values())
        ratio = max(vals) / min(vals)
        if ratio <= _MAX_RATIO: continue

        fuentes = entry.get("fuentes", {})
        tiene_cf = "maxicarrefour" in fuentes

        auditoria.append({
            "id": entry["id_unificado"],
            "nombre": entry["nombre_display"],
            "precios": precios_ok,
            "ratio": round(ratio, 2),
            "confianza": entry["confianza_match"],
            "fuentes_nombre": {k: v.get("nombre","") for k, v in fuentes.items()},
        })

        # El precio MÁS BAJO con ratio extremo casi siempre es el roto (bug de scraper)
        # Excepción: si hay solo 2 precios y ninguno es Carrefour, no podemos decidir
        precio_min_k = min(precios_ok, key=precios_ok.get)
        precio_max_k = max(precios_ok, key=precios_ok.get)

        if tiene_cf and precio_min_k != "maxicarrefour":
            # Carrefour tiene precio alto → el bajo (yaguar o mco) está roto
            entry["precios"][precio_min_k] = 0
            entry["fuentes"].pop(precio_min_k, None)
            matches_invalidos += 1
        elif tiene_cf and precio_max_k != "maxicarrefour":
            # Carrefour tiene precio bajo → el alto está roto
            entry["precios"][precio_max_k] = 0
            entry["fuentes"].pop(precio_max_k, None)
            matches_invalidos += 1
        elif "maxiconsumo" in precios_ok and ratio > 5.0:
            # Sin carrefour y ratio extremo: mco es el más probable de estar roto
            entry["precios"]["maxiconsumo"] = 0
            entry["fuentes"].pop("maxiconsumo", None)
            matches_invalidos += 1

        entry["confianza_match"] = "revisar"

    print(f"  Validacion precios: {len(auditoria)} sospechosos, {matches_invalidos} corregidos automaticamente")

    # Guardar reporte de auditoría para revisión manual
    reporte_path = os.path.join(BASE_DIR, "BRUJULA-DE-PRECIOS", "data", "processed", "auditoria_matches.json")
    with open(reporte_path, "w", encoding="utf-8") as f:
        json.dump(sorted(auditoria, key=lambda x: x["ratio"], reverse=True), f, ensure_ascii=False, indent=2)

    return list(cat.values())

# ---------------------------------------------------------------------------
def main():
    print(f"\n{'='*55}")
    print(f"  UNIFICADOR V2 — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*55}")

    print("\n[1/4] Cargando Excel...")
    yag_sku_ean, mco_sku_ean, ean_yag_sku, ean_mco_sku, ean_master, nombre_ean = cargar_excel()

    print("\n[2/4] Cargando scrapers...")
    yag  = cargar_yaguar()
    cf   = cargar_maxicarrefour()
    mco  = cargar_maxiconsumo()

    print("\n[3/4] Construyendo catálogo...")
    catalogo = construir_catalogo(yag, cf, mco, yag_sku_ean, mco_sku_ean,
                                  ean_yag_sku, ean_mco_sku, ean_master, nombre_ean)

    print("\n[4/4] Guardando...")
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(catalogo, f, ensure_ascii=False, separators=(",",":"))

    # Reporte
    total = len(catalogo)
    exactos  = [p for p in catalogo if p.get("confianza_match") == "exacto_ean"]
    f_alto   = [p for p in catalogo if p.get("confianza_match") == "fuzzy_alto"]
    f_normal = [p for p in catalogo if p.get("confianza_match") == "fuzzy"]
    unicos   = [p for p in catalogo if p.get("confianza_match") == "unico"]

    con_2 = [p for p in catalogo if sum(1 for v in p.get("precios",{}).values() if v > 0) >= 2]
    con_3 = [p for p in catalogo if sum(1 for v in p.get("precios",{}).values() if v > 0) >= 3]
    con_2_exacto = [p for p in con_2 if p.get("confianza_match") == "exacto_ean"]
    revisados    = [p for p in catalogo if p.get("confianza_match") == "revisar"]

    print(f"\n{'='*55}")
    print(f"  RESULTADO")
    print(f"{'='*55}")
    print(f"  Total productos:          {total:>6}")
    print(f"  Con 2+ precios:           {len(con_2):>6}  (antes: 3.078)")
    print(f"  Con 3 precios:            {len(con_3):>6}  (antes: 772)")
    print(f"  Con 2+ precios EXACTOS:   {len(con_2_exacto):>6}  (confianza 100%)")
    print(f"")
    print(f"  Por confianza de match:")
    print(f"    exacto_ean:  {len(exactos):>6} productos (EAN confirmado)")
    print(f"    fuzzy_alto:  {len(f_alto):>6} productos (similitud >=85%)")
    print(f"    fuzzy:       {len(f_normal):>6} productos (similitud 75-85%)")
    print(f"    revisar:     {len(revisados):>6} productos (ratio >2.5x, precio incorrecto corregido)")
    print(f"    unico:       {len(unicos):>6} productos (solo 1 mayorista)")
    print(f"")
    print(f"  Auditoria: data/processed/auditoria_matches.json")
    print(f"{'='*55}")
    print(f"  Guardado en: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
