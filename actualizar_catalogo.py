#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ACTUALIZAR CATALOGO - Brujula de Precios  v3.0
Estrategia de matching optimizada:
  1. MaxiCarrefour como hub (100% EAN) - se procesa primero
  2. CODIGOS.xlsx mapas bidireccionales: EAN -> SKU Yaguar/Maxiconsumo
  3. Maestro como fallback por nombre normalizado
  4. Selección del scraper con MÁS productos (no el más reciente)
"""

import os, json, glob, re, unicodedata
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False
    print("  [WARN] openpyxl no instalado. pip install openpyxl")

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
YAGUAR_DIR      = os.path.join(BASE_DIR, "targets", "yaguar")
MAXICARRE_DIR   = os.path.join(BASE_DIR, "targets", "maxicarrefour")
MAXICONSUMO_DIR = os.path.join(BASE_DIR, "targets", "maxiconsumo")
RAW_DIR         = os.path.join(BASE_DIR, "data", "raw")
CODIGOS_FILE    = os.path.join(RAW_DIR, "CODIGOS.xlsx")
MAESTRO_FILE    = os.path.join(RAW_DIR, "Listado Maestro 09-03.xlsx")
OUTPUT_FILE     = os.path.join(BASE_DIR, "BRUJULA-DE-PRECIOS", "data", "processed", "catalogo_unificado.json")

# ---------------------------------------------------------------------------
# Sectores
# ---------------------------------------------------------------------------
SECTOR_NORMALIZE = {
    "almacen": "Almacén", "almacén": "Almacén",
    "bebidas": "Bebidas", "bodega": "Bebidas",
    "frescos": "Frescos", "lacteos": "Frescos", "lácteos": "Frescos",
    "lácteos y productos frescos": "Frescos", "lacteos y productos frescos": "Frescos",
    "fiambreria": "Frescos", "fiambrería": "Frescos",
    "carniceria": "Frescos", "carnicería": "Frescos", "quesos": "Frescos",
    "limpieza": "Limpieza", "papeles e higiene": "Limpieza", "papeles": "Limpieza",
    "perfumeria": "Cuidado Personal", "perfumería": "Cuidado Personal",
    "cuidado personal": "Cuidado Personal",
    "bazar": "Bazar", "bazar y textil": "Bazar", "hogar y bazar": "Bazar",
    "productos de fiesta": "Bazar",
    "mascotas": "Mascotas",
    "kiosco": "Kiosco",
    "mundo bebe": "Bebés", "mundo bebé": "Bebés", "bebes": "Bebés", "bebés": "Bebés",
    "congelados": "Congelados",
    "desayuno y merienda": "Desayuno y Merienda", "desayuno": "Desayuno y Merienda",
    "panaderia": "Almacén", "panadería": "Almacén",
}

YAGUAR_SECTOR_MAP = {
    "almacen": "Almacén", "almacén": "Almacén",
    "bazar": "Bazar", "bebidas": "Bebidas", "bodega": "Bebidas",
    "desayuno": "Desayuno y Merienda", "frescos": "Frescos",
    "kiosco": "Kiosco", "limpieza": "Limpieza", "mascotas": "Mascotas",
    "papeles": "Limpieza", "perfumeria": "Cuidado Personal", "perfumería": "Cuidado Personal",
}

def normalizar_sector(raw):
    key = (raw or "").lower().strip()
    return SECTOR_NORMALIZE.get(key, (raw or "Almacén").strip().title())

def mapear_sector_yaguar(raw):
    key = (raw or "").lower().strip()
    for k, v in YAGUAR_SECTOR_MAP.items():
        if k in key:
            return v
    return normalizar_sector(raw)

# ---------------------------------------------------------------------------
# Normalización de nombres
# ---------------------------------------------------------------------------
def clave_nombre(nombre):
    """Clave de matching: sin acentos, unidades canónicas, sin puntuación."""
    n = (nombre or "").lower().strip()
    n = unicodedata.normalize("NFD", n)
    n = "".join(c for c in n if unicodedata.category(c) != "Mn")
    # Decimales con coma → punto  ("1,5" → "1.5")
    n = re.sub(r"(\d),(\d)", r"\1.\2", n)
    # Eliminar puntuación excepto punto decimal y dígitos
    n = re.sub(r"[^a-z0-9. ]", " ", n)
    # Eliminar 'x' multiplicador antes de número ("x354ml", "x 6", "x500g" → cantidad sin x)
    n = re.sub(r"\bx\s*(\d)", r"\1", n)
    # Canonicalizar unidades de volumen/masa/cantidad
    # CM3 y CC son equivalentes a ML
    n = re.sub(r"(\d+)\s*cm3\b",  lambda m: m.group(1)+"ml",  n)
    n = re.sub(r"(\d+)\s*ccm\b",  lambda m: m.group(1)+"ml",  n)
    # Litros → ml para unificar comparaciones de cantidad
    n = re.sub(r"(\d+\.?\d*)\s*lts?\b", lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    n = re.sub(r"(\d+\.?\d*)\s*lt\b",   lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    # "L" aislado después de número → ml  ("1.5 l" → "1500ml", "2 l" → "2000ml")
    n = re.sub(r"(\d+\.?\d*)\s*l\b",    lambda m: str(int(float(m.group(1))*1000))+"ml", n)
    # GRS → GR
    n = re.sub(r"(\d+)\s*grs\b", lambda m: m.group(1)+"gr", n)
    # KG → GR
    n = re.sub(r"(\d+\.?\d*)\s*kgs?\b", lambda m: str(int(float(m.group(1))*1000))+"gr", n)
    # UNIDADES: "uni" → "un" (variante larga, ej "X 2 Uni", "12 Uni")
    n = re.sub(r"(\d+)\s*uni\b", lambda m: m.group(1)+"un", n)
    # Pegar dígito+unidad (sin espacio) para matching exacto
    n = re.sub(r"(\d+)\s*(cc|ml|gr|kg|un|ul)\b", r"\1\2", n)
    # Eliminar puntos residuales
    n = re.sub(r"\.", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n

def normalizar_nombre_display(nombre):
    """Nombre limpio para mostrar al usuario."""
    n = (nombre or "").strip()
    if not n or len(n) < 3:
        return n
    # Eliminar X antes de cantidad
    n = re.sub(r"\bX\s*(?=\d)", "", n)
    # Title Case
    minusc = {"de","del","la","las","los","el","y","e","o","a","con","sin","en","al","por"}
    palabras = n.split()
    tc = []
    for i, p in enumerate(palabras):
        if not p: continue
        if i > 0 and p.lower() in minusc:
            tc.append(p.lower())
        else:
            tc.append(p[0].upper() + p[1:] if len(p) > 1 else p.upper())
    n = " ".join(tc)
    # Unidades canónicas (después del Title Case)
    n = re.sub(r"(\d)\s*[Cc][Cc]\b", r"\1 ml", n)
    n = re.sub(r"(\d)\s*[Ll][Tt][Ss]?\b", r"\1 L", n)
    n = re.sub(r"(\d)\s*[Mm][Ll][Ss]?\b", r"\1 ml", n)
    n = re.sub(r"(\d)\s*[Gg][Rr][Ss]?\b", r"\1 g", n)
    n = re.sub(r"(\d)\s*[Kk][Gg][Ss]?\b", r"\1 kg", n)
    n = re.sub(r"(\d)\s*G\b", r"\1 g", n)
    n = re.sub(r"(\d+)\s*[Uu][Nn]?\s*[Xx]\s*(\d+)\s*g\b", r"\1 u x \2 g", n)
    n = re.sub(r"(\d+)\s*[Uu][Nn]?\b", lambda m: m.group(1) + " u", n)
    # Eliminar abreviaciones de envase
    n = re.sub(r"\b(Pet|Bot|Pte|Sdo|Fco|Dsc|Brik)\b", "", n, flags=re.IGNORECASE)
    n = re.sub(r"\s+", " ", n).strip().strip(".")
    return n

def mejor_imagen(imagenes):
    """Elige la primera imagen que no sea placeholder 0000-."""
    for img in imagenes:
        if img and "/0000-" not in img:
            return img
    # Fallback: la primera que no esté vacía
    for img in imagenes:
        if img:
            return img
    return ""

# ---------------------------------------------------------------------------
# Carga de Excel
# ---------------------------------------------------------------------------
def cargar_excel_referencia():
    """
    Retorna:
      yag_sku_to_ean  : SKU Yaguar  -> EAN
      mco_sku_to_ean  : SKU Maxiconsumo -> EAN
      ean_to_yag_sku  : EAN -> SKU Yaguar   (NUEVO - para matching desde MaxiCarrefour)
      ean_to_mco_sku  : EAN -> SKU Maxiconsumo (NUEVO)
      ean_to_master   : EAN -> {nombre, sector, categoria, abc}
      nombre_norm_to_ean : nombre_normalizado -> EAN  (NUEVO - fallback por nombre)
    """
    yag_sku_to_ean  = {}
    mco_sku_to_ean  = {}
    ean_to_yag_sku  = {}
    ean_to_mco_sku  = {}
    ean_to_master   = {}
    nombre_norm_to_ean = {}

    if not EXCEL_DISPONIBLE:
        return yag_sku_to_ean, mco_sku_to_ean, ean_to_yag_sku, ean_to_mco_sku, ean_to_master, nombre_norm_to_ean

    # --- CODIGOS.xlsx ---
    if os.path.isfile(CODIGOS_FILE):
        wb = openpyxl.load_workbook(CODIGOS_FILE, read_only=True, data_only=True)

        # YAGUAR: col1=SKU Yaguar, col2=EAN
        if "YAGUAR" in wb.sheetnames:
            for row in wb["YAGUAR"].iter_rows(min_row=2, values_only=True):
                sku_raw, ean_raw = row[1], row[2]
                if not sku_raw or not ean_raw:
                    continue
                try:
                    sku = str(int(sku_raw))
                    ean = str(int(ean_raw))
                    if len(ean) >= 8:
                        yag_sku_to_ean[sku] = ean
                        ean_to_yag_sku[ean] = sku   # mapa inverso
                except (ValueError, TypeError):
                    pass

        # MAXICONSUMO: col1=SKU, col3=EAN (Código de barras)
        if "MAXICONSUMO" in wb.sheetnames:
            for row in wb["MAXICONSUMO"].iter_rows(min_row=2, values_only=True):
                sku_raw, ean_raw = row[1], row[3]
                if not sku_raw or not ean_raw:
                    continue
                try:
                    sku = str(int(sku_raw))
                    ean = str(int(ean_raw))
                    if len(ean) >= 8:
                        mco_sku_to_ean[sku] = ean
                        ean_to_mco_sku[ean] = sku   # mapa inverso
                except (ValueError, TypeError):
                    pass

        wb.close()
        print(f"  CODIGOS.xlsx: Yaguar={len(yag_sku_to_ean)} SKUs, Maxiconsumo={len(mco_sku_to_ean)} SKUs")
        print(f"  Mapas inversos: EAN->Yaguar={len(ean_to_yag_sku)}, EAN->Maxiconsumo={len(ean_to_mco_sku)}")
    else:
        print(f"  [WARN] No encontrado: {CODIGOS_FILE}")

    # --- Listado Maestro ---
    if os.path.isfile(MAESTRO_FILE):
        wb = openpyxl.load_workbook(MAESTRO_FILE, read_only=True, data_only=True)
        ws = wb["Sheet1"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre  = row[1]
            abc     = str(row[2] or "").strip().upper()
            sector  = row[3]
            familia = row[5]
            ean_col = row[6]
            barcode = row[8]
            marca   = row[9]
            categ   = row[11]

            ean_val = None
            for v in (ean_col, barcode):
                if v and str(v).strip() not in ("-", "", "None"):
                    try:
                        ean_val = str(int(float(str(v))))
                        break
                    except (ValueError, TypeError):
                        pass

            if not ean_val or not nombre:
                continue

            nombre_str = str(nombre).strip()
            ean_to_master[ean_val] = {
                "nombre":    normalizar_nombre_display(nombre_str),
                "sector":    normalizar_sector(str(sector or "")),
                "categoria": str(categ or familia or "").strip().title(),
                "marca":     str(marca or "").strip().title(),
                "abc":       abc,
            }

            # Índice por nombre normalizado -> EAN (para fallback)
            clave = clave_nombre(nombre_str)
            if clave and len(clave) > 5:
                nombre_norm_to_ean[clave] = ean_val

        wb.close()
        print(f"  Listado Maestro: {len(ean_to_master)} EANs, {len(nombre_norm_to_ean)} nombres indexados")
    else:
        print(f"  [WARN] No encontrado: {MAESTRO_FILE}")

    return yag_sku_to_ean, mco_sku_to_ean, ean_to_yag_sku, ean_to_mco_sku, ean_to_master, nombre_norm_to_ean

# ---------------------------------------------------------------------------
# Carga de scrapers
# ---------------------------------------------------------------------------
def precio_promedio(data):
    precios = [p.get("precio", 0) for p in data if p.get("precio", 0) > 0]
    return sum(precios) / len(precios) if precios else 0

def precios_validos(data):
    """Cuenta productos con precio razonable para Argentina (> $200)."""
    return sum(1 for p in data if p.get("precio", 0) > 200)

def encontrar_mejor(directorio, patron, max_check=8):
    """
    Evalua los ultimos max_check archivos y elige el mejor por:
      score = productos_con_precio_valido (>$200)
    Descarta archivos con precio promedio < $200 (bug x1000).
    """
    archivos = sorted(
        glob.glob(os.path.join(directorio, patron)),
        key=os.path.getmtime, reverse=True
    )[:max_check]

    if not archivos:
        return None, []

    mejor_archivo = None
    mejor_data    = []
    mejor_score   = -1

    for f in archivos:
        try:
            with open(f, encoding="utf-8") as fh:
                data = json.load(fh)
            prom = precio_promedio(data)
            if 0 < prom < 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 200:
                        p["precio"] = round(p["precio"] * 1000, 2)
            elif prom >= 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 100:
                        p["precio"] = round(p["precio"] * 100, 2)
            score = precios_validos(data)
            if score > mejor_score:
                mejor_score   = score
                mejor_archivo = f
                mejor_data    = data
        except Exception:
            pass

    return mejor_archivo, mejor_data


def cargar_yaguar():
    """
    Combina los últimos max_check archivos de Yaguar por SKU único.
    Para cada SKU, usa el producto del archivo más reciente con precio válido (>$200).
    Esto maximiza la cobertura de productos sin requerir scraping perfecto en cada run.
    """
    archivos = sorted(
        glob.glob(os.path.join(YAGUAR_DIR, "output_yaguar_*.json")),
        key=os.path.getmtime, reverse=True
    )[:8]

    if not archivos:
        print("  [SKIP] No se encontró output de Yaguar")
        return []

    sku_to_mejor = {}
    archivos_validos = 0

    for f in archivos:
        try:
            with open(f, encoding="utf-8") as fh:
                data = json.load(fh)
            if not data:
                continue
            prom = precio_promedio(data)
            if 0 < prom < 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 200:
                        p["precio"] = round(p["precio"] * 1000, 2)
            elif prom >= 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 100:
                        p["precio"] = round(p["precio"] * 100, 2)
            archivos_validos += 1
            for p in data:
                sku = str(p.get("sku", "")).strip()
                precio = p.get("precio", 0)
                if not sku or precio <= 0:
                    continue
                existing = sku_to_mejor.get(sku)
                if existing is None:
                    sku_to_mejor[sku] = p
                else:
                    precio_ex = existing.get("precio", 0)
                    if precio > 200 and precio_ex < 200:
                        sku_to_mejor[sku] = p
        except Exception:
            pass

    combined = list(sku_to_mejor.values())
    con_precio = precios_validos(combined)
    print(f"  Yaguar: {archivos_validos} archivos combinados -> {len(combined)} prods únicos ({con_precio} válidos)")
    return combined


def cargar_maxicarrefour():
    archivo, data = encontrar_mejor(MAXICARRE_DIR, "output_maxicarrefour_*.json")
    if not archivo:
        print("  [SKIP] No se encontró output de MaxiCarrefour")
        return []
    con_precio = sum(1 for p in data if p.get("precio", 0) > 0)
    print(f"  MaxiCarrefour: {os.path.basename(archivo)} -> {len(data)} productos ({con_precio} con precio)")
    if con_precio == 0:
        print("  [WARN] Todos los precios son 0 - cookies probablemente vencidas")
    return data


def cargar_maxiconsumo():
    """
    Carga y combina los archivos de Maxiconsumo disponibles.
    Estrategia: para cada SKU, prefiere el precio del archivo con mayor
    promedio (enriquecido) si es válido (>$200); sino usa el raw.
    Así se aprovecha la mayor cobertura del raw y la mayor calidad del enriquecido.
    """
    if not os.path.isdir(MAXICONSUMO_DIR):
        return []

    archivos = sorted(
        glob.glob(os.path.join(MAXICONSUMO_DIR, "output_maxiconsumo_*.json")),
        key=os.path.getmtime, reverse=True
    )[:8]
    if not archivos:
        return []

    # Cargar todos los archivos y construir un mapa SKU → mejor precio
    sku_to_mejor = {}   # sku → producto con mejor precio validado
    archivos_cargados = []

    for f in archivos:
        try:
            with open(f, encoding="utf-8") as fh:
                data = json.load(fh)
            prom = precio_promedio(data)
            # Fix precio × 1000 si el promedio es sospechosamente bajo
            if 0 < prom < 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 200:
                        p["precio"] = round(p["precio"] * 1000, 2)
            elif prom >= 200:
                for p in data:
                    if 0 < p.get("precio", 0) < 100:
                        p["precio"] = round(p["precio"] * 100, 2)
            archivos_cargados.append((f, data, precio_promedio(data)))
        except Exception:
            pass

    if not archivos_cargados:
        return []

    # Combinar: para cada SKU usar el archivo con precio más alto (>$200 preferido)
    for _, data, prom_f in archivos_cargados:
        for p in data:
            sku = str(p.get("sku", "")).strip()
            precio = p.get("precio", 0)
            if not sku or precio <= 0:
                continue
            existing = sku_to_mejor.get(sku)
            if existing is None:
                sku_to_mejor[sku] = p
            else:
                precio_ex = existing.get("precio", 0)
                # Preferir precio válido (>$200) sobre inválido (<$200)
                if precio > 200 and precio_ex < 200:
                    sku_to_mejor[sku] = p
                elif precio > 200 and precio_ex > 200 and precio > precio_ex:
                    # Si ambos válidos, quedarse con el del archivo más reciente
                    # (el primero en la iteración ya es el más reciente, no sobreescribir)
                    pass

    combined = list(sku_to_mejor.values())
    con_precio = sum(1 for p in combined if p.get("precio", 0) > 200)
    bajos = sum(1 for p in combined if 0 < p.get("precio", 0) < 200)
    print(f"  Maxiconsumo: {len(archivos_cargados)} archivos combinados -> {len(combined)} productos")
    print(f"    {con_precio} precios válidos (>$200), {bajos} precios bajos (<$200)")
    return combined


def cargar_hunterprice():
    ruta = os.path.join(BASE_DIR, "archive", "data_hunterprice.json")
    if not os.path.isfile(ruta):
        print("  [SKIP] No encontrado: archive/data_hunterprice.json")
        return []
    with open(ruta, encoding="utf-8") as f:
        data = json.load(f)
    # Solo los que tienen MaxiCarrefour (nuestro hub EAN) como referencia
    filtrado = [p for p in data if p.get("MAXI CARREFOUR") and p["MAXI CARREFOUR"] > 0]
    print(f"  Hunterprice: {len(data)} productos total, {len(filtrado)} con precio MaxiCarrefour")
    return filtrado

# ---------------------------------------------------------------------------
# Constructor del catálogo unificado
# ---------------------------------------------------------------------------
def construir_catalogo(yaguar_data, maxicarre_data, maxiconsumo_data,
                       yag_sku_to_ean, mco_sku_to_ean,
                       ean_to_yag_sku, ean_to_mco_sku,
                       ean_to_master, nombre_norm_to_ean):

    catalogo = {}   # prod_id -> entry

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def info_master(ean, fb_nombre, fb_sector):
        if ean and ean in ean_to_master:
            m = ean_to_master[ean]
            return m["nombre"], m["sector"], m["categoria"], m["abc"]
        return normalizar_nombre_display(fb_nombre), normalizar_sector(fb_sector), "", ""

    def nuevo_producto(prod_id, ean, nombre, imagen, sector, subcategoria, abc=""):
        return {
            "id_unificado":   prod_id,
            "ean":            ean,
            "nombre_display": nombre,
            "imagen":         imagen,
            "sector":         sector,
            "subcategoria":   subcategoria,
            "abc":            abc,
            "precios":        {"yaguar": 0, "maxicarrefour": 0, "maxiconsumo": 0},
            "fuentes":        {},
        }

    def resolver_ean(sku, sku_to_ean, nombre):
        """Obtiene EAN para un producto: CODIGOS primero, luego nombre->Maestro."""
        ean = sku_to_ean.get(str(sku).strip(), "")
        if not ean and nombre:
            ean = nombre_norm_to_ean.get(clave_nombre(nombre), "")
        return ean

    # ------------------------------------------------------------------
    # PASO 1: Indexar Yaguar y Maxiconsumo por SKU y por nombre
    # ------------------------------------------------------------------
    yag_by_sku   = {}
    yag_by_clave = {}
    for p in yaguar_data:
        sku = str(p.get("sku", "")).strip()
        nom = p.get("nombre", "")
        if sku:
            yag_by_sku[sku] = p
        if nom:
            yag_by_clave[clave_nombre(nom)] = p

    mco_by_sku   = {}
    mco_by_clave = {}
    for p in maxiconsumo_data:
        if p.get("precio", 0) <= 0:
            continue
        sku = str(p.get("sku", "")).strip()
        nom = p.get("nombre", "")
        if sku:
            mco_by_sku[sku] = p
        if nom:
            mco_by_clave[clave_nombre(nom)] = p

    yag_merged = set()   # SKUs de Yaguar ya procesados
    mco_merged = set()   # SKUs de Maxiconsumo ya procesados

    # ------------------------------------------------------------------
    # PASO 1b: Enriquecer ean_to_yag_sku y ean_to_mco_sku con Listado Maestro
    #   Tres estrategias en orden: campo ean externo → nombre exacto → Jaccard fuzzy.
    #   El índice fuzzy corre sobre nombre_norm_to_ean (25k+ entradas del Maestro)
    #   con threshold 0.60 — igual que enriquecer_eans.py pero sin depender del
    #   archivo preseleccionado por encontrar_mejor().
    # ------------------------------------------------------------------
    _FUZZ1B_TH   = 0.60
    _FUZZ1B_STOP = {"de", "la", "el", "en", "y", "x", "con", "por", "para",
                    "un", "una", "del", "los", "las", "al", "ml", "gr", "cc", "kg"}
    _fuzz_entries  = []   # (clave, word_set, ean)
    _fuzz_word_idx = defaultdict(list)
    for _clave, _ean in nombre_norm_to_ean.items():
        _ws = {w for w in _clave.split() if len(w) > 1 and w not in _FUZZ1B_STOP}
        if not _ws:
            continue
        _fi = len(_fuzz_entries)
        _fuzz_entries.append((_clave, _ws, _ean))
        for _fw in _ws:
            _fuzz_word_idx[_fw].append(_fi)

    def _fuzzy_ean_1b(nombre_prod):
        _cl   = clave_nombre(nombre_prod)
        _ws_p = {w for w in _cl.split() if len(w) > 1 and w not in _FUZZ1B_STOP}
        if not _ws_p:
            return ""
        _cands = set()
        for _w in _ws_p:
            for _i in _fuzz_word_idx.get(_w, []):
                _cands.add(_i)
        _best_sim = 0.0
        _best_ean = ""
        for _i in _cands:
            _, _ws_m, _ean = _fuzz_entries[_i]
            _inter = len(_ws_p & _ws_m)
            _union = len(_ws_p | _ws_m)
            _sim   = _inter / _union if _union else 0.0
            if _sim > _best_sim:
                _best_sim = _sim
                _best_ean = _ean
        return _best_ean if _best_sim >= _FUZZ1B_TH else ""

    ean_yag_nuevos = 0
    yag_sku_set = set(ean_to_yag_sku.values())
    for p in yaguar_data:
        sku = str(p.get("sku", "")).strip()
        if not sku or sku in yag_sku_set:
            continue
        ean_resuelto = str(p.get("ean", "") or "").strip()
        if not ean_resuelto or ean_resuelto in ("0", "None", "nan"):
            ean_resuelto = nombre_norm_to_ean.get(clave_nombre(p.get("nombre", "")), "")
        if not ean_resuelto:
            ean_resuelto = _fuzzy_ean_1b(p.get("nombre", ""))
        if ean_resuelto and ean_resuelto not in ean_to_yag_sku:
            ean_to_yag_sku[ean_resuelto] = sku
            yag_sku_set.add(sku)
            ean_yag_nuevos += 1

    ean_mco_nuevos = 0
    mco_sku_set = set(ean_to_mco_sku.values())
    for p in maxiconsumo_data:
        sku = str(p.get("sku", "")).strip()
        if not sku or sku in mco_sku_set:
            continue
        ean_resuelto = str(p.get("ean", "") or "").strip()
        if not ean_resuelto or ean_resuelto in ("0", "None", "nan"):
            ean_resuelto = nombre_norm_to_ean.get(clave_nombre(p.get("nombre", "")), "")
        if not ean_resuelto:
            ean_resuelto = _fuzzy_ean_1b(p.get("nombre", ""))
        if ean_resuelto and ean_resuelto not in ean_to_mco_sku:
            ean_to_mco_sku[ean_resuelto] = sku
            mco_sku_set.add(sku)
            ean_mco_nuevos += 1

    print(f"  Paso 1b: +{ean_yag_nuevos} EANs Yaguar via Maestro, +{ean_mco_nuevos} EANs Maxiconsumo via Maestro")

    # ------------------------------------------------------------------
    # PASO 2: MaxiCarrefour como HUB (100% EAN)
    #   Para cada producto MC busca en Yaguar y Maxiconsumo via CODIGOS
    # ------------------------------------------------------------------
    stats_mc = {"match_yag": 0, "match_mco": 0, "nuevo": 0}

    for p in maxicarre_data:
        ean    = str(p.get("ean", "")).strip()
        nombre = p.get("nombre", "")
        precio = p.get("precio", 0)
        if not ean or not nombre:
            continue

        imagen_mc   = p.get("imagen", "")
        sector_raw  = p.get("sector", "")

        nombre_display, sector, subcategoria, abc = info_master(ean, nombre, sector_raw)

        entry = nuevo_producto(ean, ean, nombre_display, imagen_mc, sector, subcategoria, abc)
        entry["precios"]["maxicarrefour"] = precio
        entry["fuentes"]["maxicarrefour"] = {"nombre": nombre, "imagen": imagen_mc}

        # Buscar Yaguar via mapa inverso EAN->SKU
        yag_sku = ean_to_yag_sku.get(ean)
        if yag_sku and yag_sku in yag_by_sku:
            yag_p = yag_by_sku[yag_sku]
            yag_precio = yag_p.get("precio", 0)
            if yag_precio > 0:
                entry["precios"]["yaguar"] = yag_precio
            entry["fuentes"]["yaguar"] = {
                "nombre": yag_p.get("nombre", ""),
                "imagen": yag_p.get("imagen", ""),
                "sku":    yag_sku,
            }
            yag_merged.add(yag_sku)
            stats_mc["match_yag"] += 1

        # Buscar Maxiconsumo via mapa inverso EAN->SKU
        mco_sku = ean_to_mco_sku.get(ean)
        if mco_sku and mco_sku in mco_by_sku:
            mco_p = mco_by_sku[mco_sku]
            mco_precio = mco_p.get("precio", 0)
            if mco_precio > 0:
                entry["precios"]["maxiconsumo"] = mco_precio
            entry["fuentes"]["maxiconsumo"] = {
                "nombre": mco_p.get("nombre", ""),
                "imagen": mco_p.get("imagen", ""),
                "sku":    mco_sku,
            }
            mco_merged.add(mco_sku)
            stats_mc["match_mco"] += 1

        # Elegir la mejor imagen (no placeholder)
        candidatas = [imagen_mc]
        if yag_sku and yag_sku in yag_by_sku:
            candidatas.append(yag_by_sku[yag_sku].get("imagen", ""))
        if mco_sku and mco_sku in mco_by_sku:
            candidatas.append(mco_by_sku[mco_sku].get("imagen", ""))
        entry["imagen"] = mejor_imagen(candidatas)

        # Si la imagen sigue siendo 0000- pero hay EAN, usar CDN Carrefour
        if "/0000-" in entry["imagen"] or not entry["imagen"]:
            entry["imagen"] = f"https://tupedido.carrefour.com.ar/imagenesPDA/{ean}.jpg"

        catalogo[ean] = entry
        stats_mc["nuevo"] += 1

    print(f"  MaxiCarrefour: {stats_mc['nuevo']} productos procesados")
    print(f"    -> Matches Yaguar via CODIGOS:      {stats_mc['match_yag']}")
    print(f"    -> Matches Maxiconsumo via CODIGOS: {stats_mc['match_mco']}")

    # ------------------------------------------------------------------
    # PASO 3: Yaguar - productos no mergeados con MaxiCarrefour
    # ------------------------------------------------------------------
    stats_yag = {"match_ean_catalogo": 0, "match_nombre_maestro": 0, "nuevo": 0}

    for p in yaguar_data:
        sku    = str(p.get("sku", "")).strip()
        nombre = p.get("nombre", "")
        precio = p.get("precio", 0)
        if not nombre or not sku:
            continue

        if sku in yag_merged:
            continue  # ya fue matcheado con MaxiCarrefour

        imagen    = p.get("imagen", "")
        sector_raw = mapear_sector_yaguar(p.get("categoria", ""))

        # Resolver EAN
        ean = resolver_ean(sku, yag_sku_to_ean, nombre)

        nombre_display, sector, subcategoria, abc = info_master(ean, nombre, sector_raw)

        if ean and ean in catalogo:
            # El EAN ya existe en catálogo (poco probable, pero por si acaso)
            catalogo[ean]["precios"]["yaguar"] = precio
            catalogo[ean]["fuentes"]["yaguar"] = {"nombre": nombre, "imagen": imagen, "sku": sku}
            if not catalogo[ean]["imagen"] or "/0000-" in catalogo[ean]["imagen"]:
                catalogo[ean]["imagen"] = mejor_imagen([imagen, catalogo[ean]["imagen"]])
            if ean in nombre_norm_to_ean.values():
                stats_yag["match_ean_catalogo"] += 1
        else:
            prod_id = ean if ean else f"yaguar_{sku}"
            if prod_id not in catalogo:
                img_final = imagen
                if "/0000-" in img_final and ean:
                    img_final = f"https://tupedido.carrefour.com.ar/imagenesPDA/{ean}.jpg"
                entry = nuevo_producto(prod_id, ean, nombre_display, img_final, sector, subcategoria, abc)
                catalogo[prod_id] = entry
                stats_yag["nuevo"] += 1

            catalogo[prod_id]["precios"]["yaguar"] = precio
            catalogo[prod_id]["fuentes"]["yaguar"] = {"nombre": nombre, "imagen": imagen, "sku": sku}
            if ean:
                stats_yag["match_nombre_maestro"] += 1

        yag_merged.add(sku)

    print(f"  Yaguar (restantes): {stats_yag['nuevo']} nuevos, "
          f"{stats_yag['match_ean_catalogo']} match EAN, "
          f"{stats_yag['match_nombre_maestro']} con EAN via Maestro")

    # ------------------------------------------------------------------
    # PASO 4: Maxiconsumo - productos no mergeados
    # ------------------------------------------------------------------
    stats_mco = {"match_ean_catalogo": 0, "match_nombre_yaguar": 0, "nuevo": 0}

    # Índice de claves de productos Yaguar sin EAN (para match por nombre)
    yag_clave_a_id = {}
    for prod_id, entry in catalogo.items():
        if not entry.get("ean") or prod_id.startswith("yaguar_"):
            clave = clave_nombre(entry["nombre_display"])
            if clave:
                yag_clave_a_id[clave] = prod_id

    for p in maxiconsumo_data:
        sku    = str(p.get("sku", "")).strip()
        nombre = p.get("nombre", "")
        precio = p.get("precio", 0)
        if not nombre or not sku or precio <= 0:
            continue

        if sku in mco_merged:
            continue  # ya matcheado con MaxiCarrefour

        imagen     = p.get("imagen", "")
        sector_raw = p.get("sector", "")

        # Resolver EAN
        ean = resolver_ean(sku, mco_sku_to_ean, nombre)

        nombre_display, sector, subcategoria, abc = info_master(ean, nombre, sector_raw)

        if ean and ean in catalogo:
            # EAN ya en catálogo
            catalogo[ean]["precios"]["maxiconsumo"] = precio
            catalogo[ean]["fuentes"]["maxiconsumo"] = {"nombre": nombre, "imagen": imagen, "sku": sku}
            if not catalogo[ean]["imagen"] or "/0000-" in catalogo[ean]["imagen"]:
                catalogo[ean]["imagen"] = mejor_imagen([imagen, catalogo[ean]["imagen"]])
            stats_mco["match_ean_catalogo"] += 1
            mco_merged.add(sku)
            continue

        # Fallback: match por nombre con Yaguar sin EAN
        clave = clave_nombre(nombre_display)
        if clave in yag_clave_a_id:
            prod_id = yag_clave_a_id[clave]
            catalogo[prod_id]["precios"]["maxiconsumo"] = precio
            catalogo[prod_id]["fuentes"]["maxiconsumo"] = {"nombre": nombre, "imagen": imagen, "sku": sku}
            if not catalogo[prod_id]["imagen"]:
                catalogo[prod_id]["imagen"] = imagen
            stats_mco["match_nombre_yaguar"] += 1
            mco_merged.add(sku)
            continue

        # Producto nuevo
        prod_id = ean if ean else f"mco_{sku}"
        if prod_id not in catalogo:
            entry = nuevo_producto(prod_id, ean, nombre_display, imagen, sector, subcategoria, abc)
            catalogo[prod_id] = entry
            stats_mco["nuevo"] += 1

        catalogo[prod_id]["precios"]["maxiconsumo"] = precio
        catalogo[prod_id]["fuentes"]["maxiconsumo"] = {"nombre": nombre, "imagen": imagen, "sku": sku}
        mco_merged.add(sku)

    print(f"  Maxiconsumo (restantes): {stats_mco['nuevo']} nuevos, "
          f"{stats_mco['match_ean_catalogo']} match EAN, "
          f"{stats_mco['match_nombre_yaguar']} match nombre Yaguar")

    # ------------------------------------------------------------------
    # PASO 5: Hunterprice bridge (triple Jaccard matching)
    #   Para cada producto de hunterprice:
    #     1. Buscar en MaxiCarrefour por nombre → obtener EAN → entry en catálogo
    #     2. Si le falta Yaguar: buscar en Yaguar scraper por nombre
    #     3. Si le falta Maxiconsumo: buscar en Maxiconsumo scraper por nombre
    #   Esto cubre productos que CODIGOS no pudo linkear por EAN.
    # ------------------------------------------------------------------
    hp_data = cargar_hunterprice()
    stats_hp = {"completados_yag": 0, "completados_mco": 0, "no_match_mc": 0}

    # Palabras de ruido para matching
    _STOP = {"x", "de", "la", "el", "y", "con", "sin", "pet", "pvc",
             "bot", "sdo", "fco", "brik", "p", "s", "en"}

    def _pals(clave):
        return {w for w in clave.split()
                if len(w) > 1 and w not in _STOP and not w.isdigit()}

    # Captura tanto números dentro de tokens de unidad ("1500ml", "500gr")
    # como números sueltos ("12", "4"). Dos grupos: uno con unidad, uno sin.
    _NUM_RE = re.compile(r"(\d+)(?:ml|gr|kg|un|cc)\b|\b(\d{2,5})\b")

    def _nums(clave):
        """Extrae números significativos (cantidad/tamaño) de una clave.
        Captura tanto '1500' de '1500ml' como números sueltos tipo '12'.
        """
        result = set()
        for m in _NUM_RE.finditer(clave):
            n = m.group(1) or m.group(2)
            if n:
                result.add(n)
        return result

    def _mejor_match(hp_ps, entries_list, word_index, threshold, hp_clave="", qty_ref=""):
        """
        Mejor match Jaccard en entries_list para hp_ps.
        Si el referente de cantidad (qty_ref > hp_clave cuando disponible) tiene números,
        el candidato debe compartir al menos uno. Esto evita cruzar tamaños distintos.
        qty_ref se usa para validar cantidad; si vacío, se cae a hp_clave.
        """
        cands = set()
        for _w in hp_ps:
            for _i in word_index.get(_w, []):
                cands.add(_i)
        if not cands:
            return None, 0.0
        # Preferir qty_ref (ej. nombre de MaxiCarrefour en catálogo) sobre hp_clave
        # para la validación de cantidad, ya que MC tiene el EAN correcto.
        _ref = qty_ref if qty_ref else hp_clave
        hp_numeros = _nums(_ref) if _ref else set()
        mejor_sim = 0.0
        mejor_val = None
        for _i in cands:
            _entry = entries_list[_i]
            _val   = _entry[0]
            _ps_c  = _entry[1]
            _clave_c = _entry[2] if len(_entry) > 2 else ""
            _inter = len(hp_ps & _ps_c)
            _union = len(hp_ps | _ps_c)
            _sim = _inter / _union if _union else 0.0
            if _sim < threshold:
                continue
            # Validar compatibilidad de cantidades
            # Si el referente tiene números y el candidato también, deben coincidir al menos 1
            if hp_numeros and _clave_c:
                _mc_numeros = _nums(_clave_c)
                if _mc_numeros and not (hp_numeros & _mc_numeros):
                    continue  # cantidades incompatibles
            if _sim > mejor_sim:
                mejor_sim = _sim
                mejor_val = _val
        if not mejor_val:
            return None, mejor_sim
        return mejor_val, mejor_sim

    _TH = 0.50  # Jaccard mínimo

    # Índice invertido MaxiCarrefour: (ean, pals, clave)
    mc_entries_hp = []
    mc_word_idx   = defaultdict(list)
    for _p in maxicarre_data:
        _ean = str(_p.get("ean", "")).strip()
        _nom = _p.get("nombre", "")
        if not _ean or not _nom:
            continue
        _cl = clave_nombre(_nom)
        _ps = _pals(_cl)
        if not _ps:
            continue
        _i = len(mc_entries_hp)
        mc_entries_hp.append((_ean, _ps, _cl))
        for _w in _ps:
            mc_word_idx[_w].append(_i)

    # Índice invertido Yaguar: (producto, pals, clave)
    yag_entries_hp = []
    yag_word_idx   = defaultdict(list)
    for _p in yaguar_data:
        _nom = _p.get("nombre", "")
        if not _nom or _p.get("precio", 0) <= 0:
            continue
        _cl = clave_nombre(_nom)
        _ps = _pals(_cl)
        if not _ps:
            continue
        _i = len(yag_entries_hp)
        yag_entries_hp.append((_p, _ps, _cl))
        for _w in _ps:
            yag_word_idx[_w].append(_i)

    # Índice invertido Maxiconsumo: (producto, pals, clave)
    mco_entries_hp = []
    mco_word_idx   = defaultdict(list)
    for _p in maxiconsumo_data:
        _nom = _p.get("nombre", "")
        if not _nom or _p.get("precio", 0) <= 0:
            continue
        _cl = clave_nombre(_nom)
        _ps = _pals(_cl)
        if not _ps:
            continue
        _i = len(mco_entries_hp)
        mco_entries_hp.append((_p, _ps, _cl))
        for _w in _ps:
            mco_word_idx[_w].append(_i)

    for hp in hp_data:
        hp_nombre = (hp.get("Descripcion_Norm") or hp.get("Nombre_Unificado") or "").strip()
        if not hp_nombre:
            continue
        hp_ps = _pals(clave_nombre(hp_nombre))
        if not hp_ps:
            continue

        hp_tiene_yag = bool(hp.get("YAGUAR"))
        hp_tiene_mco = bool(hp.get("MAXICONSUMO"))
        if not hp_tiene_yag and not hp_tiene_mco:
            continue

        hp_clave = clave_nombre(hp_nombre)

        # PASO A: Encontrar EAN via MaxiCarrefour
        ean, sim_mc = _mejor_match(hp_ps, mc_entries_hp, mc_word_idx, _TH, hp_clave)
        if not ean:
            stats_hp["no_match_mc"] += 1
            continue

        if ean not in catalogo:
            continue
        entry = catalogo[ean]

        # Cantidad de referencia = nombre del producto en MaxiCarrefour (tiene EAN correcto).
        # Usarla en PASO B y C para no cruzar tamaños distintos cuando HP no tiene unidad.
        _mc_src_nombre = entry["fuentes"].get("maxicarrefour", {}).get("nombre", "")
        _qty_ref = clave_nombre(_mc_src_nombre) if _mc_src_nombre else hp_clave

        # PASO B: Completar Yaguar si falta
        if hp_tiene_yag and entry["precios"].get("yaguar", 0) == 0:
            # Primero via CODIGOS (ya intentado en PASO 2, pero por si acaso)
            yag_sku  = ean_to_yag_sku.get(ean)
            yag_prod = yag_by_sku.get(yag_sku) if yag_sku else None
            # Si no: Jaccard sobre Yaguar scraper (usar MC como referente de cantidad)
            if not yag_prod:
                yag_prod, _ = _mejor_match(hp_ps, yag_entries_hp, yag_word_idx, _TH, hp_clave, qty_ref=_qty_ref)
                yag_sku = str(yag_prod.get("sku", "")).strip() if yag_prod else ""
            if yag_prod and yag_prod.get("precio", 0) > 0:
                entry["precios"]["yaguar"] = yag_prod["precio"]
                entry["fuentes"]["yaguar"] = {
                    "nombre": yag_prod.get("nombre", ""),
                    "imagen": yag_prod.get("imagen", ""),
                    "sku":    yag_sku,
                }
                if yag_sku:
                    yag_merged.add(yag_sku)
                stats_hp["completados_yag"] += 1

        # PASO C: Completar Maxiconsumo si falta
        if hp_tiene_mco and entry["precios"].get("maxiconsumo", 0) == 0:
            mco_sku  = ean_to_mco_sku.get(ean)
            mco_prod = mco_by_sku.get(mco_sku) if mco_sku else None
            if not mco_prod:
                mco_prod, _ = _mejor_match(hp_ps, mco_entries_hp, mco_word_idx, _TH, hp_clave, qty_ref=_qty_ref)
                mco_sku = str(mco_prod.get("sku", "")).strip() if mco_prod else ""
            if mco_prod and mco_prod.get("precio", 0) > 0:
                entry["precios"]["maxiconsumo"] = mco_prod["precio"]
                entry["fuentes"]["maxiconsumo"] = {
                    "nombre": mco_prod.get("nombre", ""),
                    "imagen": mco_prod.get("imagen", ""),
                    "sku":    mco_sku,
                }
                if mco_sku:
                    mco_merged.add(mco_sku)
                stats_hp["completados_mco"] += 1

    print(f"  Hunterprice bridge: +{stats_hp['completados_yag']} Yaguar, "
          f"+{stats_hp['completados_mco']} Maxiconsumo | "
          f"{stats_hp['no_match_mc']} sin match MC")

    # ------------------------------------------------------------------
    # PASO 6: Post-proceso
    #   - Validación cruzada de precios (descarta outliers)
    #   - Eliminar productos sin precio
    #   - Reparar imágenes 0000- con CDN Carrefour (si tienen EAN)
    #   - Fusionar duplicados de nombre exacto
    # ------------------------------------------------------------------
    lista = list(catalogo.values())

    # ------ Validación cruzada de precios ------
    # Si un producto tiene múltiples fuentes y un precio es >5x más barato
    # que el resto, se descarta como outlier (scraping error).
    # Solo aplica cuando hay al menos 2 fuentes para comparar.
    precios_descartados = 0
    for p in lista:
        precios_activos = {k: v for k, v in p["precios"].items() if v > 0}
        if len(precios_activos) < 2:
            continue
        vals = list(precios_activos.values())
        mediana = sorted(vals)[len(vals) // 2]
        for fuente, precio in list(precios_activos.items()):
            # Outlier hacia abajo: precio < mediana/4 (más de 4x más barato que la mediana)
            # No aplicar si la mediana es muy baja (producto genuinamente barato)
            if precio < mediana / 4 and mediana > 800:
                p["precios"][fuente] = 0
                if fuente in p["fuentes"]:
                    del p["fuentes"][fuente]
                precios_descartados += 1
    if precios_descartados:
        print(f"  Validación cruzada: {precios_descartados} precios outlier descartados")

    # Eliminar sin precio
    lista = [p for p in lista if any(v > 0 for v in p["precios"].values())]

    # Reparar imágenes placeholder
    for p in lista:
        img = p.get("imagen", "")
        ean = p.get("ean", "")
        if ("/0000-" in img or not img) and ean:
            p["imagen"] = f"https://tupedido.carrefour.com.ar/imagenesPDA/{ean}.jpg"

    def _fusionar_grupo(items):
        """Fusiona una lista de productos al mejor representante (prioridad: EAN real)."""
        base = max(items, key=lambda x: bool(x.get("ean")))
        for item in items:
            for fuente, precio in item["precios"].items():
                if precio > 0 and base["precios"].get(fuente, 0) == 0:
                    base["precios"][fuente] = precio
            for fuente, info in item.get("fuentes", {}).items():
                if fuente not in base["fuentes"]:
                    base["fuentes"][fuente] = info
            if not base.get("imagen") or "/0000-" in base.get("imagen", ""):
                if item.get("imagen") and "/0000-" not in item.get("imagen", ""):
                    base["imagen"] = item["imagen"]
            if not base.get("abc") and item.get("abc"):
                base["abc"] = item["abc"]
        return base

    def _es_sintetico(prod_id):
        return str(prod_id).startswith("yaguar_") or str(prod_id).startswith("mco_")

    # Paso 6a: Fusionar duplicados de nombre_display exacto
    por_nombre = defaultdict(list)
    for p in lista:
        por_nombre[p["nombre_display"]].append(p)

    lista_paso6a = []
    for nombre, items in por_nombre.items():
        if len(items) == 1:
            lista_paso6a.append(items[0])
        else:
            lista_paso6a.append(_fusionar_grupo(items))

    # Paso 6b: Fusionar duplicados por nombre normalizado
    #   Solo fusiona si al menos uno tiene ID sintético (yaguar_XXX / mco_XXX),
    #   lo que indica que una fuente no pudo resolver el EAN del mismo producto.
    #   Si ambos tienen EAN real, son SKUs genuinamente distintos → no tocar.
    por_clave = defaultdict(list)
    for p in lista_paso6a:
        por_clave[clave_nombre(p["nombre_display"])].append(p)

    lista_final = []
    fusiones_norm = 0
    for clave_norm, items in por_clave.items():
        if len(items) == 1:
            lista_final.append(items[0])
            continue
        hay_sintetico = any(_es_sintetico(p["id_unificado"]) for p in items)
        if not hay_sintetico:
            # Todos tienen EAN real → SKUs genuinamente distintos, mantener separados
            lista_final.extend(items)
        else:
            lista_final.append(_fusionar_grupo(items))
            fusiones_norm += 1

    if fusiones_norm:
        print(f"  Paso 6b: {fusiones_norm} duplicados por nombre normalizado fusionados")

    # ------------------------------------------------------------------
    # PASO 6c: Fusión fuzzy de productos complementarios
    #   Para cada producto con precios faltantes, busca en los productos
    #   que tienen esa(s) fuente(s) faltante(s) usando Jaccard > 0.82.
    #   Prioridad de base: maxicarrefour (tiene EAN) > yaguar > maxiconsumo.
    # ------------------------------------------------------------------
    _STOP6 = {"de", "la", "el", "y", "con", "sin", "pet", "pvc",
              "bot", "sdo", "fco", "brik", "p", "s", "en"}
    # Mismo patrón que _NUM_RE: captura números dentro de unidades y sueltos
    _NUM6  = re.compile(r"(\d+)(?:ml|gr|kg|un|cc)\b|\b(\d{2,6})\b")

    def _w6(clave):
        return {w for w in clave.split() if len(w) > 1 and w not in _STOP6 and not w.isdigit()}

    def _n6(clave):
        result = set()
        for m in _NUM6.finditer(clave):
            n = m.group(1) or m.group(2)
            if n:
                result.add(n)
        return result

    _TH6 = 0.65

    def _buscar_candidato(ws_p, ns_p, index_entries, index_wi, usados):
        """Devuelve (idx_en_lista_final, sim) del mejor match fuzzy."""
        cands = set()
        for w in ws_p:
            for ei in index_wi.get(w, []):
                cands.add(ei)
        mejor_sim = 0.0
        mejor_idx = None
        for ei in cands:
            lf_idx, ws_c, cl_c, ns_c = index_entries[ei]
            if lf_idx in usados:
                continue
            inter = len(ws_p & ws_c)
            union = len(ws_p | ws_c)
            sim   = inter / union if union else 0.0
            if sim < _TH6:
                continue
            if ns_p and ns_c and not (ns_p & ns_c):
                continue  # cantidades incompatibles
            if sim > mejor_sim:
                mejor_sim = sim
                mejor_idx = lf_idx
        return mejor_idx, mejor_sim

    # Construir índices por fuente
    def _build_index(fuente, lista):
        """(idx_lista_final, ws, cl, ns) para productos que TIENEN fuente y solo fuente."""
        entries = []
        wi      = defaultdict(list)
        for idx, p in enumerate(lista):
            if p["precios"].get(fuente, 0) <= 0:
                continue
            cl = clave_nombre(p["nombre_display"])
            ws = _w6(cl)
            ns = _n6(cl)
            if not ws:
                continue
            ei = len(entries)
            entries.append((idx, ws, cl, ns))
            for w in ws:
                wi[w].append(ei)
        return entries, wi

    mc_idx_entries,  mc_idx_wi  = _build_index("maxicarrefour", lista_final)
    yag_idx_entries, yag_idx_wi = _build_index("yaguar", lista_final)
    mco_idx_entries, mco_idx_wi = _build_index("maxiconsumo", lista_final)

    fusiones_fuzzy = 0
    usados_como_base = set()   # índices de lista_final que ya absorbieron algo

    # parche: { idx_eliminado: idx_base }
    parches = {}

    for idx_p, p in enumerate(lista_final):
        if idx_p in usados_como_base or idx_p in parches:
            continue
        pr = p["precios"]
        tiene_mc  = pr.get("maxicarrefour", 0) > 0
        tiene_yag = pr.get("yaguar", 0) > 0
        tiene_mco = pr.get("maxiconsumo", 0) > 0
        n_fuentes = sum([tiene_mc, tiene_yag, tiene_mco])
        if n_fuentes == 3:
            continue  # completo

        cl_p = clave_nombre(p["nombre_display"])
        ws_p = _w6(cl_p)
        ns_p = _n6(cl_p)
        if not ws_p:
            continue

        # Buscar fuentes faltantes en orden de prioridad
        for fuente_falt, entries_f, wi_f in [
            ("maxicarrefour", mc_idx_entries,  mc_idx_wi),
            ("yaguar",        yag_idx_entries, yag_idx_wi),
            ("maxiconsumo",   mco_idx_entries, mco_idx_wi),
        ]:
            if pr.get(fuente_falt, 0) > 0:
                continue  # ya tiene esta fuente

            lf_idx, sim = _buscar_candidato(ws_p, ns_p, entries_f, wi_f, usados_como_base | set(parches.keys()) | {idx_p})
            if lf_idx is None:
                continue

            p_cand = lista_final[lf_idx]
            # Solo absorber si el candidato tiene SOLO esa fuente (o pocas fuentes)
            # → evitar partir un producto ya bien matcheado
            n_cand = sum(1 for v in p_cand["precios"].values() if v > 0)
            if n_cand > 2:
                continue  # candidato ya tiene muchos precios, no arriesgar

            # ¿Cuál es la "base" (el que tiene EAN)?
            p_tiene_ean = bool(p.get("ean"))
            cand_tiene_ean = bool(p_cand.get("ean"))
            if cand_tiene_ean and not p_tiene_ean:
                # Cand es la base, p se funde en cand
                parches[idx_p] = lf_idx
                usados_como_base.add(lf_idx)
            elif p_tiene_ean:
                # p es la base, cand se funde en p
                parches[lf_idx] = idx_p
                usados_como_base.add(idx_p)
            else:
                # Ninguno tiene EAN: base = el que tiene más fuentes
                if n_cand >= n_fuentes:
                    parches[idx_p] = lf_idx
                    usados_como_base.add(lf_idx)
                else:
                    parches[lf_idx] = idx_p
                    usados_como_base.add(idx_p)
            fusiones_fuzzy += 1
            break  # una fusión por producto es suficiente

    # Aplicar parches
    for idx_elim, idx_base in parches.items():
        p_elim = lista_final[idx_elim]
        p_base = lista_final[idx_base]
        for fuente, precio in p_elim["precios"].items():
            if precio > 0 and p_base["precios"].get(fuente, 0) == 0:
                p_base["precios"][fuente] = precio
        for fuente, info in p_elim.get("fuentes", {}).items():
            if fuente not in p_base["fuentes"]:
                p_base["fuentes"][fuente] = info
        if not p_base.get("abc") and p_elim.get("abc"):
            p_base["abc"] = p_elim["abc"]
        p_elim["_eliminar"] = True

    lista_final = [p for p in lista_final if not p.get("_eliminar")]

    print(f"  Paso 6c: {fusiones_fuzzy} fusiones fuzzy complementarias")

    # ------------------------------------------------------------------
    # PASO 6d: Validación de cantidad entre fuentes (cleanup defensivo)
    #   Para cada producto con 2+ fuentes, extrae la cantidad (en unidades
    #   canónicas) del nombre de cada fuente. Si una fuente tiene cantidades
    #   incompatibles con el "ancla" (MC > Yaguar > MCO), la elimina.
    #   Sólo actúa cuando hay diferencia > 2x para evitar falsos positivos
    #   en variantes con nombres levemente distintos (ej. 950ml vs 930ml).
    # ------------------------------------------------------------------
    _QTY_RE = re.compile(r"(\d+)(?:ml|gr|kg|un|cc)\b|\b(\d{2,5})\b")

    def _src_nums(nombre):
        """Extrae numeros de cantidad del nombre crudo de una fuente."""
        cl = clave_nombre(nombre)
        result = set()
        for m in _QTY_RE.finditer(cl):
            n = m.group(1) or m.group(2)
            if n:
                result.add(int(n))
        return result

    fuentes_eliminadas_6d = 0
    _ANCHOR_ORDER = ["maxicarrefour", "yaguar", "maxiconsumo"]

    for p in lista_final:
        precios_activos = {k: v for k, v in p["precios"].items() if v > 0}
        if len(precios_activos) < 2:
            continue
        fuentes = p.get("fuentes", {})

        # Elegir el ancla: la fuente con mayor confianza (MC > Yaguar > MCO)
        ancla = None
        for _f in _ANCHOR_ORDER:
            if precios_activos.get(_f, 0) > 0:
                ancla = _f
                break
        if not ancla:
            continue

        ancla_nombre = fuentes.get(ancla, {}).get("nombre", "")
        if not ancla_nombre:
            continue
        ancla_nums = _src_nums(ancla_nombre)
        if not ancla_nums:
            continue  # ancla sin info de cantidad: no validar

        ancla_max = max(ancla_nums)

        for fuente in list(precios_activos.keys()):
            if fuente == ancla:
                continue
            src_nombre = fuentes.get(fuente, {}).get("nombre", "")
            if not src_nombre:
                continue
            src_nums = _src_nums(src_nombre)
            if not src_nums:
                continue  # fuente sin cantidad: no sancionar
            # Si no comparten ningún número Y la diferencia máxima es > 2x → mismatch
            if not (ancla_nums & src_nums):
                src_max = max(src_nums)
                ratio = max(ancla_max, src_max) / max(1, min(ancla_max, src_max))
                if ratio >= 2.0:
                    p["precios"][fuente] = 0
                    if fuente in p["fuentes"]:
                        del p["fuentes"][fuente]
                    fuentes_eliminadas_6d += 1

    if fuentes_eliminadas_6d:
        print(f"  Paso 6d: {fuentes_eliminadas_6d} fuentes con cantidad incompatible eliminadas")

    # Eliminar productos que quedaron sin precio tras la limpieza 6d
    lista_final = [p for p in lista_final if any(v > 0 for v in p["precios"].values())]

    return lista_final


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("ACTUALIZADOR DE CATALOGO v3.0 - Brujula de Precios")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    print("\nCargando tablas de referencia (Excel)...")
    (yag_sku_to_ean, mco_sku_to_ean,
     ean_to_yag_sku, ean_to_mco_sku,
     ean_to_master, nombre_norm_to_ean) = cargar_excel_referencia()

    print("\nCargando datos de scrapers (mejor archivo por cantidad)...")
    yaguar      = cargar_yaguar()
    maxicarre   = cargar_maxicarrefour()
    maxiconsumo = cargar_maxiconsumo()

    print("\nConstruyendo catálogo unificado...")
    catalogo = construir_catalogo(
        yaguar, maxicarre, maxiconsumo,
        yag_sku_to_ean, mco_sku_to_ean,
        ean_to_yag_sku, ean_to_mco_sku,
        ean_to_master, nombre_norm_to_ean,
    )

    # Stats
    con_yag  = sum(1 for p in catalogo if p["precios"]["yaguar"] > 0)
    con_mc   = sum(1 for p in catalogo if p["precios"]["maxicarrefour"] > 0)
    con_mco  = sum(1 for p in catalogo if p["precios"]["maxiconsumo"] > 0)
    multi    = sum(1 for p in catalogo if sum(1 for v in p["precios"].values() if v > 0) >= 2)
    tres     = sum(1 for p in catalogo if sum(1 for v in p["precios"].values() if v > 0) == 3)
    abc_a_multi = sum(1 for p in catalogo if p.get("abc") == "A" and
                      sum(1 for v in p["precios"].values() if v > 0) >= 2)
    sin_img  = sum(1 for p in catalogo if not p.get("imagen"))

    print(f"\n{'='*60}")
    print(f"RESULTADO FINAL")
    print(f"{'='*60}")
    print(f"  Total productos con precio:   {len(catalogo)}")
    print(f"  Con precio Yaguar:            {con_yag}")
    print(f"  Con precio MaxiCarrefour:     {con_mc}")
    print(f"  Con precio Maxiconsumo:       {con_mco}")
    print(f"  Con 2+ precios (comparativa): {multi}")
    print(f"  Con 3 precios:                {tres}")
    print(f"  ABC=A con 2+ precios:         {abc_a_multi}")
    print(f"  Sin imagen:                   {sin_img}")

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(catalogo, f, ensure_ascii=False, indent=2)

    print(f"\n  Guardado en: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
